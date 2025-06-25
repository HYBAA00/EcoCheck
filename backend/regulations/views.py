from django.shortcuts import render
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import HttpResponse
import csv
import json
from django.template.loader import render_to_string
from django.contrib.auth import get_user_model

# Import conditionnel pour Excel
try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .models import (
    TreatmentType, Law, Regulation, FeeStructure, ValidationCycle,
    SystemConfiguration, AuditLog, SystemMetrics, AdminNotification
)
from .serializers import (
    TreatmentTypeSerializer, LawSerializer, RegulationSerializer,
    FeeStructureSerializer, ValidationCycleSerializer, SystemConfigurationSerializer,
    AuditLogSerializer, SystemMetricsSerializer, AdminDashboardStatsSerializer,
    UserManagementSerializer, ExportDataSerializer, AdminNotificationSerializer
)
from accounts.models import User, CompanyProfile, Employee, Authority, Administrator
from certifications.models import CertificationRequest, Payment, Certificate

User = get_user_model()

class IsAdminPermission(permissions.BasePermission):
    """Permission personnalisée pour vérifier si l'utilisateur est admin"""
    
    def has_permission(self, request, view):
        return (
            request.user and 
            request.user.is_authenticated and 
            request.user.role == 'admin'
        )

class TreatmentTypeViewSet(viewsets.ModelViewSet):
    queryset = TreatmentType.objects.all()
    serializer_class = TreatmentTypeSerializer
    permission_classes = [IsAdminPermission]
    
    def get_queryset(self):
        queryset = TreatmentType.objects.all()
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        return queryset
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Statistiques des types de traitement"""
        active_count = TreatmentType.objects.filter(is_active=True).count()
        inactive_count = TreatmentType.objects.filter(is_active=False).count()
        
        # Statistiques par type de traitement
        treatment_stats = []
        for treatment_type in TreatmentType.objects.filter(is_active=True):
            requests_count = CertificationRequest.objects.filter(
                treatment_type=treatment_type.name
            ).count()
            treatment_stats.append({
                'name': treatment_type.name,
                'requests_count': requests_count,
                'fee': float(treatment_type.certification_fee)
            })
        
        return Response({
            'active_count': active_count,
            'inactive_count': inactive_count,
            'treatment_stats': treatment_stats
        })

class LawViewSet(viewsets.ModelViewSet):
    queryset = Law.objects.all()
    serializer_class = LawSerializer
    permission_classes = [IsAdminPermission]

    def get_queryset(self):
        queryset = Law.objects.all()
        category = self.request.query_params.get('category', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if category:
            queryset = queryset.filter(category=category)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
            
        return queryset.order_by('-effective_date')

class RegulationViewSet(viewsets.ModelViewSet):
    queryset = Regulation.objects.all()
    serializer_class = RegulationSerializer
    permission_classes = [IsAdminPermission]

class FeeStructureViewSet(viewsets.ModelViewSet):
    queryset = FeeStructure.objects.all()
    serializer_class = FeeStructureSerializer
    permission_classes = [IsAdminPermission]

    def get_queryset(self):
        queryset = FeeStructure.objects.all()
        treatment_type = self.request.query_params.get('treatment_type', None)
        is_active = self.request.query_params.get('is_active', None)
        
        if treatment_type:
            queryset = queryset.filter(treatment_type_id=treatment_type)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
            
        return queryset.order_by('-effective_from')

class ValidationCycleViewSet(viewsets.ModelViewSet):
    queryset = ValidationCycle.objects.all()
    serializer_class = ValidationCycleSerializer
    permission_classes = [IsAdminPermission]
    
    @action(detail=True, methods=['post'])
    def set_as_default(self, request, pk=None):
        """Définir un cycle comme cycle par défaut"""
        cycle = self.get_object()
        
        # Retirer le statut par défaut des autres cycles du même type de traitement
        ValidationCycle.objects.filter(
            treatment_type=cycle.treatment_type,
            is_default=True
        ).update(is_default=False)
        
        # Définir ce cycle comme par défaut
        cycle.is_default = True
        cycle.save()
        
        return Response({'message': 'Cycle défini comme par défaut'})

class SystemConfigurationViewSet(viewsets.ModelViewSet):
    queryset = SystemConfiguration.objects.all()
    serializer_class = SystemConfigurationSerializer
    permission_classes = [IsAdminPermission]

    def get_queryset(self):
        queryset = SystemConfiguration.objects.all()
        category = self.request.query_params.get('category', None)
        
        if category:
            queryset = queryset.filter(category=category)
            
        return queryset.order_by('category', 'name')

    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Liste des catégories de configuration"""
        categories = SystemConfiguration.objects.values_list('category', flat=True).distinct()
        return Response(list(categories))

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AuditLog.objects.all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAdminPermission]
    
    def get_queryset(self):
        queryset = AuditLog.objects.all()
        
        # Filtres
        user_id = self.request.query_params.get('user', None)
        action = self.request.query_params.get('action', None)
        date_from = self.request.query_params.get('date_from', None)
        date_to = self.request.query_params.get('date_to', None)
        success = self.request.query_params.get('success', None)
        
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        if action:
            queryset = queryset.filter(action=action)
        if success is not None:
            queryset = queryset.filter(success=success.lower() == 'true')
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)
            
        return queryset.order_by('-timestamp')

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Statistiques des logs d'audit pour la page Reports"""
        today = timezone.now().date()
        last_24h = timezone.now() - timedelta(hours=24)
        
        # Statistiques de base
        total_actions = AuditLog.objects.count()
        successful_actions = AuditLog.objects.filter(success=True).count()
        
        # Utilisateurs actifs (qui ont fait des actions dans les dernières 24h)
        unique_users = AuditLog.objects.filter(
            timestamp__gte=last_24h
        ).values('user').distinct().count()
        
        # Sessions actives (approximation basée sur les connexions récentes)
        recent_logins = AuditLog.objects.filter(
            action='login',
            timestamp__gte=last_24h,
            success=True
        ).count()
        
        # Données réelles du système
        active_users_total = User.objects.filter(is_active=True).count()
        certificates_issued = Certificate.objects.count()
        pending_requests = CertificationRequest.objects.filter(
            status__in=['submitted', 'under_review']
        ).count()
        
        stats = {
            'total_actions': total_actions,
            'successful_actions': successful_actions,
            'unique_users': unique_users or active_users_total,  # Fallback aux utilisateurs actifs
            'active_sessions': recent_logins or 0,
            'certificates_issued': certificates_issued,
            'pending_requests': pending_requests,
            'success_rate': (successful_actions / total_actions * 100) if total_actions > 0 else 0,
        }
        
        # Actions les plus fréquentes
        top_actions = AuditLog.objects.values('action').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        stats['top_actions'] = list(top_actions)
        
        return Response(stats)

class SystemMetricsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SystemMetrics.objects.all()
    serializer_class = SystemMetricsSerializer
    permission_classes = [IsAdminPermission]
    
    @action(detail=False, methods=['post'])
    def generate_daily_metrics(self, request):
        """Générer les métriques quotidiennes"""
        today = timezone.now().date()
        
        # Vérifier si les métriques d'aujourd'hui existent déjà
        if SystemMetrics.objects.filter(date=today).exists():
            return Response({'message': 'Métriques déjà générées pour aujourd\'hui'})
        
        # Calculer les métriques
        metrics = SystemMetrics.objects.create(
            date=today,
            total_requests=CertificationRequest.objects.count(),
            pending_requests=CertificationRequest.objects.filter(
                status__in=['submitted', 'under_review']
            ).count(),
            approved_requests=CertificationRequest.objects.filter(status='approved').count(),
            rejected_requests=CertificationRequest.objects.filter(status='rejected').count(),
            
            total_payments=Payment.objects.filter(status='completed').aggregate(
                total=Sum('total_amount')
            )['total'] or 0,
            pending_payments=Payment.objects.filter(status='pending').count(),
            completed_payments=Payment.objects.filter(status='completed').count(),
            
            total_users=User.objects.count(),
            active_users=User.objects.filter(is_active=True).count(),
            new_registrations=User.objects.filter(date_joined__date=today).count(),
            
            certificates_issued=Certificate.objects.count(),
            certificates_expired=Certificate.objects.filter(expiry_date__lt=today).count(),
            
            avg_processing_time=0,  # À calculer selon la logique métier
            avg_approval_rate=0,  # À calculer selon la logique métier
        )
        
        return Response(SystemMetricsSerializer(metrics).data)

class AdminDashboardViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminPermission]

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Statistiques principales du dashboard admin"""
        
        # Statistiques des utilisateurs
        total_users = User.objects.count()
        total_enterprises = User.objects.filter(role='enterprise').count()
        total_employees = User.objects.filter(role='employee').count()
        total_authorities = User.objects.filter(role='authority').count()
        
        # Statistiques des demandes
        total_requests = CertificationRequest.objects.count()
        pending_requests = CertificationRequest.objects.filter(
            status__in=['submitted', 'under_review']
        ).count()
        approved_requests = CertificationRequest.objects.filter(status='approved').count()
        rejected_requests = CertificationRequest.objects.filter(status='rejected').count()
        
        # Statistiques des paiements
        payments_stats = Payment.objects.filter(status='completed').aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        total_payments = payments_stats['total'] or 0
        completed_payments = payments_stats['count']
        pending_payments = Payment.objects.filter(status='pending').aggregate(
            total=Sum('total_amount'),
            count=Count('id')
        )
        
        # Statistiques des certificats
        certificates_issued = Certificate.objects.count()
        today = timezone.now().date()
        certificates_expired = Certificate.objects.filter(expiry_date__lt=today).count()
        
        # Activités récentes
        recent_activities = []
        recent_logs = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]
        for log in recent_logs:
            recent_activities.append({
                'id': log.id,
                'action': log.get_action_display(),
                'description': log.description,
                'user': log.user.username if log.user else 'Système',
                'timestamp': log.timestamp.isoformat(),
                'success': log.success
            })
        
        stats_data = {
            'total_users': total_users,
            'total_enterprises': total_enterprises,
            'total_employees': total_employees,
            'total_authorities': total_authorities,
            'total_requests': total_requests,
            'pending_requests': pending_requests,
            'approved_requests': approved_requests,
            'rejected_requests': rejected_requests,
            'total_payments': total_payments,
            'pending_payments': pending_payments['total'] or 0,
            'completed_payments': total_payments,
            'certificates_issued': certificates_issued,
            'certificates_expired': certificates_expired,
            'recent_activities': recent_activities
        }
        
        serializer = AdminDashboardStatsSerializer(data=stats_data)
        if serializer.is_valid():
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UserManagementViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserManagementSerializer
    permission_classes = [IsAdminPermission]
    
    def get_queryset(self):
        queryset = User.objects.all()
        
        # Filtres
        role = self.request.query_params.get('role', None)
        is_active = self.request.query_params.get('is_active', None)
        search = self.request.query_params.get('search', None)
        
        if role:
            queryset = queryset.filter(role=role)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
            
        return queryset.order_by('-date_joined')
    
    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activer/désactiver un utilisateur"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        
        # Log de l'action
        AuditLog.objects.create(
            action='update',
            description=f'Utilisateur {"activé" if user.is_active else "désactivé"}',
            user=request.user,
            content_type='User',
            object_id=user.id,
            object_repr=str(user)
        )
        
        return Response({
            'message': f'Utilisateur {"activé" if user.is_active else "désactivé"}',
            'is_active': user.is_active
        })
    
    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Réinitialiser le mot de passe d'un utilisateur"""
        user = self.get_object()
        
        # Générer un nouveau mot de passe temporaire
        import secrets
        import string
        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) 
                              for _ in range(12))
        
        user.set_password(temp_password)
        user.save()
        
        # Log de l'action
        AuditLog.objects.create(
            action='update',
            description='Mot de passe réinitialisé',
            user=request.user,
            content_type='User',
            object_id=user.id,
            object_repr=str(user)
        )
        
        return Response({
            'message': 'Mot de passe réinitialisé',
            'temp_password': temp_password
        })

class DataExportViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminPermission]
    
    @action(detail=False, methods=['post'])
    def export(self, request):
        """Exporter des données selon les paramètres"""
        serializer = ExportDataSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        export_type = serializer.validated_data['export_type']
        date_from = serializer.validated_data.get('date_from')
        date_to = serializer.validated_data.get('date_to')
        format_type = serializer.validated_data['format']
        filters = serializer.validated_data.get('filters', {})
        
        # Déterminer les données à exporter
        if export_type == 'users':
            data = self._export_users(date_from, date_to, filters)
            filename = f'users_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        elif export_type == 'requests':
            data = self._export_requests(date_from, date_to, filters)
            filename = f'requests_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        elif export_type == 'payments':
            data = self._export_payments(date_from, date_to, filters)
            filename = f'payments_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        elif export_type == 'certificates':
            data = self._export_certificates(date_from, date_to, filters)
            filename = f'certificates_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        elif export_type == 'audit_logs':
            data = self._export_audit_logs(date_from, date_to, filters)
            filename = f'audit_logs_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
        else:
            return Response({'error': 'Type d\'export non supporté'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Générer le fichier selon le format
        if format_type == 'csv':
            response = self._generate_csv(data, filename)
        elif format_type == 'excel':
            response = self._generate_excel(data, filename)
        else:
            return Response({'error': 'Format non supporté'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Log de l'export
        AuditLog.objects.create(
            action='export',
            description=f'Export {export_type} en format {format_type}',
            user=request.user,
            additional_data={
                'export_type': export_type,
                'format': format_type,
                'date_from': str(date_from) if date_from else None,
                'date_to': str(date_to) if date_to else None,
                'filters': filters
            }
        )
        
        return response
    
    def _export_users(self, date_from, date_to, filters):
        queryset = User.objects.all()
        if date_from:
            queryset = queryset.filter(date_joined__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date_joined__date__lte=date_to)
        
        return [
            {
                'ID': user.id,
                'Nom d\'utilisateur': user.username,
                'Email': user.email,
                'Prénom': user.first_name,
                'Nom': user.last_name,
                'Rôle': user.get_role_display(),
                'Téléphone': user.phone,
                'Actif': 'Oui' if user.is_active else 'Non',
                'Date d\'inscription': user.date_joined.strftime('%d/%m/%Y %H:%M'),
                'Dernière connexion': user.last_login.strftime('%d/%m/%Y %H:%M') if user.last_login else 'Jamais'
            }
            for user in queryset
        ]
    
    def _export_requests(self, date_from, date_to, filters):
        queryset = CertificationRequest.objects.select_related('company').all()
        if date_from:
            queryset = queryset.filter(submission_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(submission_date__lte=date_to)
        
        return [
            {
                'ID': req.id,
                'Entreprise': req.company.business_name,
                'ICE': req.company.ice_number,
                'Type de traitement': req.treatment_type,
                'Date de soumission': req.submission_date.strftime('%d/%m/%Y'),
                'Statut': req.get_status_display(),
                'Assigné à': req.assigned_to.user.username if req.assigned_to else '',
                'Validé par': req.validated_by.user.username if req.validated_by else '',
                'Révisé par': req.reviewed_by.username if req.reviewed_by else ''
            }
            for req in queryset
        ]
    
    def _export_payments(self, date_from, date_to, filters):
        queryset = Payment.objects.select_related('certification_request__company').all()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        
        return [
            {
                'ID': payment.id,
                'Entreprise': payment.certification_request.company.business_name,
                'Demande ID': payment.certification_request.id,
                'Montant': float(payment.amount),
                'Frais': float(payment.fees),
                'Total': float(payment.total_amount),
                'Méthode': payment.get_payment_method_display(),
                'Statut': payment.get_status_display(),
                'Date de création': payment.created_at.strftime('%d/%m/%Y %H:%M'),
                'Date de paiement': payment.payment_date.strftime('%d/%m/%Y %H:%M') if payment.payment_date else ''
            }
            for payment in queryset
        ]
    
    def _export_certificates(self, date_from, date_to, filters):
        queryset = Certificate.objects.select_related('certification_request__company').all()
        if date_from:
            queryset = queryset.filter(issue_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(issue_date__lte=date_to)
        
        return [
            {
                'ID': cert.id,
                'Numéro': cert.number,
                'Entreprise': cert.certification_request.company.business_name,
                'Type de traitement': cert.treatment_type,
                'Date d\'émission': cert.issue_date.strftime('%d/%m/%Y'),
                'Date d\'expiration': cert.expiry_date.strftime('%d/%m/%Y'),
                'Statut': cert.status,
                'Demande ID': cert.certification_request.id
            }
            for cert in queryset
        ]
    
    def _export_audit_logs(self, date_from, date_to, filters):
        queryset = AuditLog.objects.select_related('user').all()
        if date_from:
            queryset = queryset.filter(timestamp__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(timestamp__date__lte=date_to)
        
        return [
            {
                'ID': log.id,
                'Action': log.get_action_display(),
                'Description': log.description,
                'Utilisateur': log.user.username if log.user else 'Système',
                'Horodatage': log.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
                'Adresse IP': log.ip_address or '',
                'Type d\'objet': log.content_type or '',
                'ID d\'objet': log.object_id or '',
                'Succès': 'Oui' if log.success else 'Non',
                'Message d\'erreur': log.error_message or ''
            }
            for log in queryset
        ]
    
    def _generate_csv(self, data, filename):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return response
    
    def _generate_excel(self, data, filename):
        if not EXCEL_AVAILABLE:
            return HttpResponse(
                'Excel export non disponible. Veuillez installer openpyxl.',
                status=500
            )
        
        wb = Workbook()
        ws = wb.active
        
        if data:
            # En-têtes
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Données
            for row, item in enumerate(data, 2):
                for col, value in enumerate(item.values(), 1):
                    ws.cell(row=row, column=col, value=value)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

class AdminNotificationViewSet(viewsets.ModelViewSet):
    """ViewSet pour les notifications admin"""
    serializer_class = AdminNotificationSerializer
    permission_classes = [IsAdminPermission]

    def get_queryset(self):
        user = self.request.user
        
        # Filtrer par utilisateur admin
        if user.role == 'admin':
            # Récupérer les notifications pour cet admin spécifique + les notifications générales
            return AdminNotification.objects.filter(
                Q(recipient=user) | Q(recipient__isnull=True)
            ).select_related('recipient')
        
        return AdminNotification.objects.none()
    
    @action(detail=True, methods=['post'])
    def mark_as_read(self, request, pk=None):
        """Marquer une notification comme lue"""
        notification = self.get_object()
        notification.mark_as_read()
        
        # Log de l'action
        AuditLog.objects.create(
            action='view',
            description=f'Notification marquée comme lue: {notification.title}',
            user=request.user,
            content_type='AdminNotification',
            object_id=notification.id
        )
        
        return Response({'message': 'Notification marquée comme lue'})
    
    @action(detail=True, methods=['post'])
    def dismiss(self, request, pk=None):
        """Ignorer une notification"""
        notification = self.get_object()
        notification.dismiss()
        
        # Log de l'action
        AuditLog.objects.create(
            action='update',
            description=f'Notification ignorée: {notification.title}',
            user=request.user,
            content_type='AdminNotification',
            object_id=notification.id
        )
        
        return Response({'message': 'Notification ignorée'})
    
    @action(detail=False, methods=['post'])
    def mark_all_as_read(self, request):
        """Marquer toutes les notifications comme lues"""
        notifications = self.get_queryset().filter(is_read=False)
        count = notifications.count()
        
        for notification in notifications:
            notification.mark_as_read()
        
        # Log de l'action
        AuditLog.objects.create(
            action='update',
            description=f'{count} notifications marquées comme lues',
            user=request.user,
            additional_data={'count': count}
        )
        
        return Response({'message': f'{count} notifications marquées comme lues'})

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """Compter les notifications non lues"""
        count = self.get_queryset().filter(is_read=False, is_dismissed=False).count()
        return Response({'unread_count': count})

    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Récupérer les notifications récentes (dernières 24h)"""
        from datetime import timedelta
        recent_time = timezone.now() - timedelta(hours=24)
        
        notifications = self.get_queryset().filter(
            created_at__gte=recent_time,
            is_dismissed=False
        ).order_by('-created_at')[:10]
        
        serializer = self.get_serializer(notifications, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def create_notification(self, request):
        """Créer une nouvelle notification (pour tests ou admin)"""
        data = request.data
        
        notification = AdminNotification.create_notification(
            title=data.get('title', 'Notification'),
            message=data.get('message', ''),
            notification_type=data.get('type', 'system_alert'),
            priority=data.get('priority', 'medium'),
            recipient=request.user if data.get('for_me') else None,
            action_url=data.get('action_url'),
            action_label=data.get('action_label'),
            metadata=data.get('metadata', {})
        )
        
        # Log de la création
        AuditLog.objects.create(
            action='create',
            description=f'Notification créée: {notification.title}',
            user=request.user,
            content_type='AdminNotification',
            object_id=notification.id
        )
        
        serializer = self.get_serializer(notification)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
